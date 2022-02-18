import scipy.io
import math
import numpy as np
from matplotlib import pyplot as plt
import matplotlib.patches as patches
import matplotlib.transforms as mt
from agent import Agent, cal_interior_cost, cal_group_cost
from tools.utility import get_central_vertices, smooth_cv
import pandas as pd

from openpyxl import load_workbook

illustration_needed = False
print_needed = False

# load data
mat = scipy.io.loadmat('./data/NDS_data.mat')
# full interaction information
inter_info = mat['interaction_info']
'''
inter_info:
0-1: [position x] [position y]
2-3: [velocity x] [velocity y]
4: [heading]
5: [velocity overall = sqrt(vx^2+xy^2)]
6: [curvature] (only for left-turn vehicles)
dt = 0.12s 
'''
# the number of go-straight vehicles that interact with the left-turn vehicle
inter_num = mat['interact_agent_num']

virtual_agent_IPV_range = np.array([-4, -3, -2, -1, 0, 1, 2, 3, 4]) * math.pi / 9


def draw_rectangle(x, y, deg):

    car_len = 5
    car_wid = 2
    fig = plt.figure()
    ax = fig.add_subplot(111)

    r1 = patches.Rectangle((x-car_wid/2, y-car_len/2), car_wid, car_len, color="blue", alpha=0.50)
    r2 = patches.Rectangle((x-car_wid/2, y-car_len/2), car_wid, car_len, color="red", alpha=0.50)

    t2 = mt.Affine2D().rotate_deg_around(x, y, deg) + ax.transData
    r2.set_transform(t2)

    ax.add_patch(r1)
    ax.add_patch(r2)

    plt.grid(True)
    plt.axis('equal')

    plt.show()


def visualize_nds(case_id):
    # abstract interaction info. of a given case
    case_info = inter_info[case_id]
    # left-turn vehicle
    lt_info = case_info[0]
    # go-straight vehicles
    gs_info_multi = case_info[1:inter_num[0, case_id] + 1]

    fig = plt.figure(1)
    manager = plt.get_current_fig_manager()
    manager.full_screen_toggle()
    ax1 = fig.add_subplot(121)
    ax2 = fig.add_subplot(122)
    ax2.set(xlim=[-22, 53], ylim=[-31, 57])
    img = plt.imread('background_pic/Jianhexianxia.jpg')
    ax2.imshow(img, extent=[-22, 53, -31, 57])

    for t in range(np.size(lt_info, 0)):
        t_end = t + 6
        ax1.cla()
        ax1.set(xlim=[-22, 53], ylim=[-31, 57])
        img = plt.imread('background_pic/Jianhexianxia.jpg')
        ax1.imshow(img, extent=[-22, 53, -31, 57])

        # position of go-straight vehicles
        for gs_id in range(np.size(gs_info_multi, 0)):
            if np.size(gs_info_multi[gs_id], 0) > t and not gs_info_multi[gs_id][t, 0] == 0:
                # position
                ax1.scatter(gs_info_multi[gs_id][t, 0], gs_info_multi[gs_id][t, 1],
                            s=120,
                            alpha=0.9,
                            color='red',
                            label='go-straight')
                # future track
                t_end_gs = min(t + 6, np.size(gs_info_multi[gs_id], 0))
                ax1.plot(gs_info_multi[gs_id][t:t_end_gs, 0], gs_info_multi[gs_id][t:t_end_gs, 1],
                         alpha=0.8,
                         color='red')

        # position of left-turn vehicle
        ax1.scatter(lt_info[t, 0], lt_info[t, 1],
                    s=120,
                    alpha=0.9,
                    color='blue',
                    label='left-turn')
        # future track
        ax1.plot(lt_info[t:t_end, 0], lt_info[t:t_end, 1],
                 alpha=0.8,
                 color='blue')
        # ax1.legend()
        plt.pause(0.1)

    # show full track of all agents
    ax2.plot(lt_info[:, 0], lt_info[:, 1],
             alpha=0.8,
             color='blue')
    for gs_id in range(np.size(gs_info_multi, 0)):
        # find solid frames
        frames = np.where(gs_info_multi[gs_id][:, 0] < 1e-3)
        # the first solid frame id
        frame_start = len(frames[0])
        # tracks
        ax2.plot(gs_info_multi[gs_id][frame_start:, 0], gs_info_multi[gs_id][frame_start:, 1],
                 alpha=0.8,
                 color='red')
    plt.show()


def find_inter_od(case_id):
    """
    find the starting and end frame of each FC agent that interacts with LT agent
    :param case_id:
    :return:
    """
    case_info = inter_info[case_id]
    lt_info = case_info[0]

    # find co-present gs agents (not necessarily interacting)
    gs_info_multi = case_info[1:inter_num[0, case_id] + 1]

    # find interacting gs agent
    init_id = 0
    inter_o = np.zeros(np.size(gs_info_multi, 0))
    inter_d = np.zeros(np.size(gs_info_multi, 0))
    for i in range(np.size(gs_info_multi, 0)):
        gs_agent_temp = gs_info_multi[i]
        solid_frame = np.nonzero(gs_agent_temp[:, 0])[0]
        solid_range = range(solid_frame[0], solid_frame[-1])
        inter_frame = solid_frame[0] + np.array(
            np.where(gs_agent_temp[solid_range, 1] - lt_info[solid_range, 1] < 0)[0])

        # find start and end frame with each gs agent
        if inter_frame.size > 1:
            if i == init_id:
                inter_o[i] = inter_frame[0]
            else:
                inter_o[i] = max(inter_frame[0], inter_d[i - 1])
            inter_d[i] = max(inter_frame[-1], inter_d[i - 1])
        else:
            init_id += 1
    return inter_o, inter_d


def analyze_nds(case_id):
    """
    estimate IPV in natural driving data and write results into excels
    :param case_id:
    :return:
    """
    inter_o, inter_d = find_inter_od(case_id)
    case_info = inter_info[case_id]
    lt_info = case_info[0]

    # find co-present gs agents (not necessarily interacting)
    gs_info_multi = case_info[1:inter_num[0, case_id] + 1]

    # identify IPV
    start_time = 0
    ipv_collection = np.zeros_like(lt_info[:, 0:2])
    ipv_error_collection = np.ones_like(lt_info[:, 0:2])

    # set figure
    fig = plt.figure(1)
    ax1 = fig.add_subplot(121)
    ax2 = fig.add_subplot(122)

    inter_id = 0
    inter_id_save = inter_id
    file_name = './outputs/NDS_analysis/' + str(case_id) + '.xlsx'
    df = pd.DataFrame()
    df.to_excel(file_name)

    for t in range(np.size(lt_info, 0)):

        "find current interacting agent"
        flag = 0
        for i in range(np.size(gs_info_multi, 0)):
            if inter_o[i] <= t < inter_d[i]:  # switch to next interacting agent
                # update interaction info
                flag = 1
                inter_id = i
                if print_needed:
                    print('inter_id', inter_id)
                start_time = max(int(inter_o[inter_id]), t - 10)

        # save data of last one
        if inter_id_save < inter_id or t == inter_d[-1]:
            if inter_d[inter_id_save] - inter_o[inter_id_save] > 3:
                '''
                inter_id_save < inter_id：  interacting agent changed
                t == inter_d[-1]:  end frame of the last agent
                inter_d[inter_id_save]-inter_o[inter_id_save] > 3：  interacting period is long enough
                '''

                # print('t:', t)
                # print('inter_id:', inter_id)
                # print('inter_id_save:', inter_id_save)
                book = load_workbook(file_name)
                df_ipv = pd.DataFrame(ipv_collection[int(inter_o[inter_id_save]) + 3: int(inter_d[inter_id_save]), :])
                df_ipv_error = pd.DataFrame(
                    ipv_error_collection[int(inter_o[inter_id_save]) + 3: int(inter_d[inter_id_save]), :])

                with pd.ExcelWriter(file_name) as writer:
                    if 'Sheet1' not in book.sheetnames:
                        writer.book = book
                    df_ipv.to_excel(writer, startcol=0, index=False, sheet_name=str(inter_id_save))
                    df_ipv_error.to_excel(writer, startcol=2, index=False, sheet_name=str(inter_id_save))
                    writer.save()
            inter_id_save = inter_id

        "IPV estimation process"
        if flag and (t - start_time > 3):

            "====simulation-based method===="
            # generate two agents
            init_position_lt = lt_info[start_time, 0:2]
            init_velocity_lt = lt_info[start_time, 2:4]
            init_heading_lt = lt_info[start_time, 4]
            agent_lt = Agent(init_position_lt, init_velocity_lt, init_heading_lt, 'lt_nds')
            lt_track = lt_info[start_time:t + 1, 0:2]

            init_position_gs = gs_info_multi[inter_id][start_time, 0:2]
            init_velocity_gs = gs_info_multi[inter_id][start_time, 2:4]
            init_heading_gs = gs_info_multi[inter_id][start_time, 4]
            agent_gs = Agent(init_position_gs, init_velocity_gs, init_heading_gs, 'gs_nds')
            gs_track = gs_info_multi[inter_id][start_time:t + 1, 0:2]

            # estimate ipv
            agent_lt.estimate_self_ipv_in_NDS(lt_track, gs_track)
            ipv_collection[t, 0] = agent_lt.ipv
            ipv_error_collection[t, 0] = agent_lt.ipv_error

            agent_gs.estimate_self_ipv_in_NDS(gs_track, lt_track)
            ipv_collection[t, 1] = agent_gs.ipv
            ipv_error_collection[t, 1] = agent_gs.ipv_error

            if print_needed:
                print('left turn', agent_lt.ipv, agent_lt.ipv_error)
                print('go straight', agent_gs.ipv, agent_gs.ipv_error)
            "====end of simulation-based method===="

            "====cost-based method===="
            # # load observed trajectories
            # lt_track_observed = lt_info[start_time:t + 1, 0:2]
            # gs_track_observed = gs_info_multi[inter_id][start_time:t + 1, 0:2]
            #
            # # generate two agents
            # init_position_lt = lt_info[start_time, 0:2]
            # init_velocity_lt = lt_info[start_time, 2:4]
            # init_heading_lt = lt_info[start_time, 4]
            # agent_lt = Agent(init_position_lt, init_velocity_lt, init_heading_lt, 'lt_nds')
            # agent_lt.ipv = 0
            # init_position_gs = gs_info_multi[inter_id][start_time, 0:2]
            # init_velocity_gs = gs_info_multi[inter_id][start_time, 2:4]
            # init_heading_gs = gs_info_multi[inter_id][start_time, 4]
            # agent_gs = Agent(init_position_gs, init_velocity_gs, init_heading_gs, 'gs_nds')
            # agent_gs.ipv = 0
            #
            # # plan under selfish assumption
            # lt_track_selfish = agent_lt.solve_game_IBR(gs_track_observed)
            # lt_track_selfish = lt_track_selfish[:, 0:2]
            # gs_track_selfish = agent_lt.solve_game_IBR(lt_track_observed)
            # gs_track_selfish = gs_track_selfish[:, 0:2]
            #
            # # cost results in observation
            # lt_interior_cost_observed = cal_interior_cost([], lt_track_observed, 'lt_nds')
            # gs_interior_cost_observed = cal_interior_cost([], gs_track_observed, 'gs_nds')
            # group_cost_observed = cal_group_cost([lt_track_observed, gs_track_observed])
            #
            # # cost result in assumption
            # lt_interior_cost_assumed = cal_interior_cost([], lt_track_selfish, 'lt_nds')
            # gs_interior_cost_assumed = cal_interior_cost([], gs_track_selfish, 'gs_nds')
            # group_cost_lt_assumed = cal_group_cost([lt_track_selfish, gs_track_observed])
            # group_cost_gs_assumed = cal_group_cost([lt_track_observed, gs_track_selfish])

            # ipv_collection[t, 0] =
            # ipv_collection[t, 1] =
            "====end of cost-based method===="

            "illustration"
            if illustration_needed:
                ax1.cla()
                ax1.set(ylim=[-2, 2])

                x_range = range(max(0, t - 10), t)
                smoothed_ipv_lt, _ = smooth_cv(np.array([x_range, ipv_collection[x_range, 0]]).T)
                smoothed_ipv_error_lt, _ = smooth_cv(np.array([x_range, ipv_error_collection[x_range, 0]]).T)
                smoothed_x = smoothed_ipv_lt[:, 0]
                # plot ipv
                ax1.plot(smoothed_x, smoothed_ipv_lt[:, 1], 'blue')
                # plot error bar
                ax1.fill_between(smoothed_x, smoothed_ipv_lt[:, 1] - smoothed_ipv_error_lt[:, 1],
                                 smoothed_ipv_lt[:, 1] + smoothed_ipv_error_lt[:, 1],
                                 alpha=0.4,
                                 color='blue',
                                 label='estimated lt IPV')

                smoothed_ipv_gs, _ = smooth_cv(np.array([x_range, ipv_collection[x_range, 1]]).T)
                smoothed_ipv_error_gs, _ = smooth_cv(np.array([x_range, ipv_error_collection[x_range, 1]]).T)
                # plot ipv
                ax1.plot(smoothed_x, smoothed_ipv_gs[:, 1], 'red')
                # plot error bar
                ax1.fill_between(smoothed_x, smoothed_ipv_gs[:, 1] - smoothed_ipv_error_gs[:, 1],
                                 smoothed_ipv_gs[:, 1] + smoothed_ipv_error_gs[:, 1],
                                 alpha=0.4,
                                 color='red',
                                 label='estimated gs IPV')
                ax1.legend()

                # show trajectory and plans
                ax2.cla()
                ax2.set(xlim=[-22, 53], ylim=[-31, 57])
                img = plt.imread('background_pic/Jianhexianxia.jpg')
                ax2.imshow(img, extent=[-22, 53, -31, 57])
                cv1, _ = get_central_vertices('lt_nds', [lt_info[start_time, 0], lt_info[start_time, 1]])
                cv2, _ = get_central_vertices('gs_nds', [gs_info_multi[inter_id][start_time, 0],
                                                         gs_info_multi[inter_id][start_time, 1]])
                ax2.plot(cv1[:, 0], cv1[:, 1])
                ax2.plot(cv2[:, 0], cv2[:, 1])

                # actual track
                ax2.scatter(lt_info[start_time:t, 0], lt_info[start_time:t, 1],
                            s=50,
                            alpha=0.5,
                            color='blue',
                            label='left-turn')
                candidates_lt = agent_lt.virtual_track_collection
                for track_lt in candidates_lt:
                    ax2.plot(track_lt[:, 0], track_lt[:, 1], color='green', alpha=0.5)
                ax2.scatter(gs_info_multi[inter_id][start_time:t, 0], gs_info_multi[inter_id][start_time:t, 1],
                            s=50,
                            alpha=0.5,
                            color='red',
                            label='go-straight')
                candidates_gs = agent_gs.virtual_track_collection
                for track_gs in candidates_gs:
                    ax2.plot(track_gs[:, 0], track_gs[:, 1], color='green', alpha=0.5)
                ax2.legend()

                plt.pause(0.3)

        elif inter_id is None:
            if print_needed:
                print('no interaction')

        elif t - start_time < 3:
            if print_needed:
                print('no results, more observation needed')


def analyze_ipv_in_nds(case_id, fig=False):
    file_name = './outputs/NDS_analysis/v2/' + str(case_id) + '.xlsx'
    file = pd.ExcelFile(file_name)
    num_sheet = len(file.sheet_names)
    # print(num_sheet)
    start_x = 0
    value_lt_plotted = []
    error_lt_plotted = []
    value_gs_plotted = []
    error_gs_plotted = []
    for i in range(num_sheet):
        "get ipv data from excel"
        df_ipv_data = pd.read_excel(file_name, sheet_name=i)
        ipv_data_temp = df_ipv_data.values
        ipv_value = ipv_data_temp[:, 0:2]
        ipv_error = ipv_data_temp[:, 2:]

        "draw ipv value and error bar"

        if fig:
            x = start_x + np.arange(len(ipv_value[:, 0]))
            start_x = start_x + len(ipv_value[:, 0]) + 3
            print(start_x)

            if len(x) > 3:
                # left turn
                smoothed_ipv_value_lt, _ = smooth_cv(np.array([x, ipv_value[:, 0]]).T)
                smoothed_ipv_error_lt, _ = smooth_cv(np.array([x, ipv_error[:, 0]]).T)
                plt.plot(smoothed_ipv_value_lt[:, 0], smoothed_ipv_value_lt[:, 1],
                         color='blue')
                plt.fill_between(smoothed_ipv_value_lt[:, 0], smoothed_ipv_value_lt[:, 1] - smoothed_ipv_error_lt[:, 1],
                                 smoothed_ipv_value_lt[:, 1] + smoothed_ipv_error_lt[:, 1],
                                 alpha=0.4,
                                 color='blue')

                # go straight
                smoothed_ipv_value_gs, _ = smooth_cv(np.array([x, ipv_value[:, 1]]).T)
                smoothed_ipv_error_gs, _ = smooth_cv(np.array([x, ipv_error[:, 1]]).T)
                plt.plot(smoothed_ipv_value_gs[:, 0], smoothed_ipv_value_gs[:, 1],
                         color='red')
                plt.fill_between(smoothed_ipv_value_gs[:, 0], smoothed_ipv_value_gs[:, 1] - smoothed_ipv_error_gs[:, 1],
                                 smoothed_ipv_value_gs[:, 1] + smoothed_ipv_error_gs[:, 1],
                                 alpha=0.4,
                                 color='red')

                # # save plotted data
                # value_lt_plotted.append(smoothed_ipv_value_lt)
                # value_gs_plotted.append(smoothed_ipv_value_gs)
                # error_lt_plotted.append(smoothed_ipv_error_lt)
                # error_gs_plotted.append(smoothed_ipv_error_gs)

            else:  # too short to be fitted
                # left turn
                plt.plot(x, ipv_value[:, 0],
                         color='red')
                plt.fill_between(x, ipv_value[:, 0] - ipv_error[:, 0],
                                 ipv_value[:, 0] + ipv_error[:, 0],
                                 alpha=0.4,
                                 color='red',
                                 label='estimated lt IPV')

                # go straight
                plt.plot(x, ipv_value[:, 1],
                         color='blue')
                plt.fill_between(x, ipv_value[:, 1] - ipv_error[:, 1],
                                 ipv_value[:, 1] + ipv_error[:, 1],
                                 alpha=0.4,
                                 color='blue',
                                 label='estimated gs IPV')
                # # save plotted data
                # value_lt_plotted.append(ipv_value[:, 0])
                # value_gs_plotted.append(ipv_value[:, 1])
                # error_lt_plotted.append(ipv_error[:, 0])
                # error_gs_plotted.append(ipv_value[:, 1])
            # plt.pause(1)
    plt.show()

    "select crossing event"
    inter_o, inter_d = find_inter_od(case_id)

    invalid_list = np.where(inter_d - inter_o < 4)
    index = list(invalid_list[0])
    inter_o = np.delete(inter_o, index)
    inter_d = np.delete(inter_d, index)

    case_info = inter_info[case_id]
    lt_info = case_info[0]
    crossing_frame = -1
    if max(lt_info[:, 0] > 20):
        crossing_frame = np.min(np.where(lt_info[:, 0] > 20))  # use the x position of left turn agent

    # find the interacting go-straight agent
    crossing_id = -1
    for i in range(len(inter_o)):
        if inter_o[i] < crossing_frame < inter_d[i]:
            crossing_id = i
            break

    # save ipv during the crossing event
    ipv_data_crossing = []
    ipv_data_non_crossing = []
    if not crossing_id == -1:
        df_ipv_data = pd.read_excel(file_name, sheet_name=crossing_id)
        ipv_data_crossing = df_ipv_data.values[1:, :]

    for sheet_id in range(num_sheet):
        if not sheet_id == crossing_id:
            df_ipv_data = pd.read_excel(file_name, sheet_name=sheet_id)
            ipv_data_non_crossing.append(df_ipv_data.values[1:, :])

    return crossing_id, ipv_data_crossing, ipv_data_non_crossing


def show_ipv_distribution():
    ipv_cross_lt = []
    ipv_cross_gs = []
    ipv_non_cross_lt = []
    ipv_non_cross_gs = []
    for i in range(np.size(inter_info, 0)):
        # print(i)
        _, ipv_cross_temp, ipv_non_cross_temp = analyze_ipv_in_nds(i, False)
        if len(ipv_cross_temp) > 0:
            ipv_cross_lt.append(ipv_cross_temp[:, 0])
            ipv_cross_gs.append(ipv_cross_temp[:, 1])
        if len(ipv_non_cross_temp) > 0:
            for idx in range(len(ipv_non_cross_temp)):
                # print(ipv_non_cross[idx][:, 0])
                ipv_non_cross_lt.append(ipv_non_cross_temp[idx][:, 0])
                ipv_non_cross_gs.append(ipv_non_cross_temp[idx][:, 1])

    "calculate mean ipv value of each type"
    mean_ipv_cross_lt = np.array([np.mean(ipv_cross_lt[0])])
    mean_ipv_cross_gs = np.array([np.mean(ipv_cross_gs[0])])
    mean_ipv_non_cross_lt = np.array([np.mean(ipv_non_cross_lt[0])])
    mean_ipv_non_cross_gs = np.array([np.mean(ipv_non_cross_gs[0])])
    for i in range(len(ipv_cross_lt) - 1):
        if np.size(ipv_cross_lt[i + 1], 0) > 4:
            mean_temp1 = np.array([np.mean(ipv_cross_lt[i + 1])])
            mean_ipv_cross_lt = np.concatenate((mean_ipv_cross_lt, mean_temp1), axis=0)
    for i in range(len(ipv_cross_gs) - 1):
        if np.size(ipv_cross_gs[i + 1], 0) > 4:
            mean_temp2 = np.array([np.mean(ipv_cross_gs[i + 1])])
            mean_ipv_cross_gs = np.concatenate((mean_ipv_cross_gs, mean_temp2), axis=0)
    for i in range(len(ipv_non_cross_lt) - 1):
        if np.size(ipv_non_cross_lt[i + 1], 0) > 4:
            mean_temp3 = np.array([np.mean(ipv_non_cross_lt[i + 1])])
            mean_ipv_non_cross_lt = np.concatenate((mean_ipv_non_cross_lt, mean_temp3), axis=0)
    for i in range(len(ipv_non_cross_gs) - 1):
        if np.size(ipv_non_cross_gs[i + 1], 0) > 4:
            mean_temp4 = np.array([np.mean(ipv_non_cross_gs[i + 1])])
            mean_ipv_non_cross_gs = np.concatenate((mean_ipv_non_cross_gs, mean_temp4), axis=0)

    filename = './outputs/ipv_distribution.xlsx'
    with pd.ExcelWriter(filename) as writer:

        data1 = np.vstack((mean_ipv_cross_gs, mean_ipv_cross_lt))
        df_ipv_distribution = pd.DataFrame(data1.T)
        df_ipv_distribution.to_excel(writer, startcol=0, index=False)

        data2 = np.vstack((mean_ipv_non_cross_gs, mean_ipv_non_cross_lt))
        df_ipv_distribution = pd.DataFrame(data2.T)
        df_ipv_distribution.to_excel(writer, startcol=2, index=False)

    plt.figure(1)
    plt.title('Left-turn vehicle rushed')
    plt.hist(mean_ipv_cross_lt,
             alpha=0.5,
             color='blue',
             label='left-turn vehicle')
    plt.hist(mean_ipv_cross_gs,
             alpha=0.5,
             color='red',
             label='go-straight vehicle')
    plt.legend()
    plt.xlabel('IPV')
    plt.ylabel('Counts')

    plt.figure(2)
    plt.title('Left-turn vehicle yielded')
    plt.hist(mean_ipv_non_cross_lt,
             alpha=0.5,
             color='blue',
             label='left-turn vehicle')
    plt.hist(mean_ipv_non_cross_gs,
             alpha=0.5,
             color='red',
             label='go-straight vehicle')
    plt.legend()
    plt.xlabel('IPV')
    plt.ylabel('Counts')
    plt.show()


if __name__ == '__main__':
    "analyze ipv in NDS"
    # analyze_nds(61)

    "show trajectories in NDS"
    visualize_nds(30)
    #
    # cross_id, ipv_data_cross, ipv_data_non_cross = analyze_ipv_in_nds(30, True)

    # o, d = find_inter_od(29)

    # draw_rectangle(5, 5, 45)

    # show_ipv_distribution()


